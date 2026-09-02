#!/usr/bin/env python3
"""
excel_jira.py — Excel ↔ Jira CLI integration
  1. Bulk create Jira stories from an Excel file
  2. Sync Jira data back to Excel (new sheet)
  3. Generate a blank Excel template to fill in
"""
import json
import os
import sys
from datetime import datetime, timedelta

import inquirer
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from requests.auth import HTTPBasicAuth
from rich.console import Console
from rich.table import Table

console = Console()

JIRA_BASE_URL = "https://maersk-tools.atlassian.net"
JIRA_PROJECT  = "MID1"

BOARDS = {
    "WS4 - AI/ML":        41741,
    "MIDAS - Cloud Gov":  43666,
    "MIDAS - SMUS (K)":   47357,
    "WS1 - DBX":          40984,
    "WS2 - FEP Decking":  47952,
}

# Expected columns in the upload sheet (case-insensitive)
EXPECTED_COLS = ["summary", "parent_issue", "description", "priority", "story_points", "assignee_email"]

STORY_POINTS_MAP = {
    "0": 0, "1": 1, "2": 2, "3": 3, "5": 5, "8": 8, "13": 13,
    "not set": 0, "trivial": 1, "small": 2, "medium": 3,
    "large": 5, "very large": 8, "too big": 13,
}

# ── Credentials ───────────────────────────────────────────────────────────────

def get_credentials():
    creds_file = os.path.expanduser("~/.jira_cli_creds")
    email, token = "", ""

    if os.path.exists(creds_file):
        with open(creds_file) as f:
            data = json.load(f)
            email = data.get("email", "")
            token = data.get("token", "")
        console.print(f"\n[green]✅ Using saved credentials for:[/green] {email}")
        use_saved = inquirer.prompt([
            inquirer.Confirm("use", message="Use saved credentials?", default=True)
        ])
        if not use_saved or not use_saved["use"]:
            email, token = "", ""

    if not email or not token:
        answers = inquirer.prompt([
            inquirer.Text("email", message="Jira Email"),
            inquirer.Password("token", message="Jira API Token"),
        ])
        email = answers["email"]
        token = answers["token"]
        save  = inquirer.prompt([inquirer.Confirm("save", message="Remember credentials?", default=True)])
        if save and save["save"]:
            with open(creds_file, "w") as f:
                json.dump({"email": email, "token": token}, f)
            os.chmod(creds_file, 0o600)
            console.print("[green]✅ Credentials saved.[/green]")

    return email, token


def validate_credentials(email, token):
    resp = requests.get(
        f"{JIRA_BASE_URL}/rest/api/3/myself",
        auth=HTTPBasicAuth(email, token),
        headers={"Accept": "application/json"},
        timeout=10,
    )
    if resp.status_code == 200:
        console.print(f"[green]✅ Authenticated as:[/green] {resp.json().get('displayName', '')}")
        return True
    console.print(f"[red]❌ Authentication failed ({resp.status_code}).[/red]")
    return False


def auth(email, token):
    return HTTPBasicAuth(email, token)


# ── Story Points Field ─────────────────────────────────────────────────────────

def adf_to_text(node):
    """Flatten an Atlassian Document Format (ADF) body into plain text."""
    if not node:
        return ""
    parts = []

    def walk(n):
        if isinstance(n, dict):
            if n.get("type") == "text":
                parts.append(n.get("text", ""))
            for child in n.get("content", []) or []:
                walk(child)
        elif isinstance(n, list):
            for item in n:
                walk(item)

    walk(node)
    return " ".join(p for p in parts if p).strip()


def fetch_issue_comments(email, token, key):
    resp = requests.get(
        f"{JIRA_BASE_URL}/rest/api/3/issue/{key}/comment",
        auth=auth(email, token),
        headers={"Accept": "application/json"},
        timeout=15,
    )
    if resp.status_code != 200:
        return []
    out = []
    for c in resp.json().get("comments", []):
        out.append({
            "Key":     key,
            "Author":  (c.get("author") or {}).get("displayName", ""),
            "Created": (c.get("created") or "")[:10],
            "Text":    adf_to_text(c.get("body")),
        })
    return out


def fetch_issue_changelog(email, token, key):
    resp = requests.get(
        f"{JIRA_BASE_URL}/rest/api/3/issue/{key}",
        params={"expand": "changelog"},
        auth=auth(email, token),
        headers={"Accept": "application/json"},
        timeout=15,
    )
    if resp.status_code != 200:
        return []
    out = []
    for h in resp.json().get("changelog", {}).get("histories", []):
        author  = (h.get("author") or {}).get("displayName", "")
        created = (h.get("created") or "")[:10]
        for item in h.get("items", []):
            out.append({
                "Key":     key,
                "Author":  author,
                "Created": created,
                "Field":   item.get("field", ""),
                "From":    item.get("fromString") or "",
                "To":      item.get("toString") or "",
            })
    return out


def get_story_points_field(email, token):
    resp = requests.get(
        f"{JIRA_BASE_URL}/rest/api/3/field",
        auth=auth(email, token),
        headers={"Accept": "application/json"},
        timeout=10,
    )
    if resp.status_code != 200:
        return None
    for field in resp.json():
        if "story point" in field.get("name", "").lower():
            return field["id"]
    return None


# ── Generate Template ─────────────────────────────────────────────────────────

def generate_template():
    console.rule("[bold blue]📄 Generate Excel Template[/bold blue]")
    ans = inquirer.prompt([
        inquirer.Text("filename", message="Save template as", default="jira_stories_template.xlsx")
    ])
    filename = ans["filename"]
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Stories"

    headers = ["Summary *", "Parent Issue", "Description", "Priority", "Story Points", "Assignee Email"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 30

    # Sample rows
    samples = [
        ["As a user, I want to login via SSO", "MID1-100", "Implement SSO login flow using SAML", "High", "5", "user@example.com"],
        ["Fix null pointer in payment service", "",         "NPE thrown when cart is empty",        "Highest", "3", ""],
        ["Update README with setup steps",      "",         "Add docker-compose and env var docs",  "Low",     "1", ""],
    ]
    note_fill = PatternFill("solid", fgColor="FFF2CC")
    for row_data in samples:
        ws.append(row_data)
        for col in range(1, 7):
            ws.cell(row=ws.max_row, column=col).fill = note_fill

    # Notes row
    ws.append([])
    ws.append(["📌 Notes:"])
    ws.append(["* Summary is required. All other columns are optional."])
    ws.append(["* Priority values: Highest, High, Medium, Low, Lowest"])
    ws.append(["* Story Points values: 0, 1, 2, 3, 5, 8, 13"])

    wb.save(filename)
    console.print(f"[green]✅ Template saved:[/green] {filename}")
    console.print("[dim]Fill in the Stories sheet and run 'Bulk Create from Excel'.[/dim]")


# ── Bulk Create from Excel ────────────────────────────────────────────────────

def bulk_create_from_excel(email, token):
    console.rule("[bold blue]📥 Bulk Create Stories from Excel[/bold blue]")

    ans = inquirer.prompt([
        inquirer.Text("file", message="Path to Excel file", default="jira_stories_template.xlsx"),
        inquirer.List("board", message="Select Board", choices=list(BOARDS.keys())),
    ])

    filepath = ans["file"]
    if not os.path.exists(filepath):
        console.print(f"[red]❌ File not found: {filepath}[/red]")
        return

    try:
        wb = load_workbook(filepath)
    except Exception as e:
        console.print(f"[red]❌ Could not open file: {e}[/red]")
        return

    # Pick sheet
    sheet_names = wb.sheetnames
    sheet_ans = inquirer.prompt([
        inquirer.List("sheet", message="Select sheet", choices=sheet_names)
    ])
    ws = wb[sheet_ans["sheet"]]

    # Parse headers
    headers = [str(ws.cell(row=1, column=c).value or "").strip().lower().replace(" *", "").replace(" ", "_")
               for c in range(1, ws.max_column + 1)]

    def col_idx(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    idx = {col: col_idx(col) for col in EXPECTED_COLS}

    if idx["summary"] is None:
        console.print("[red]❌ 'Summary' column not found. Please use the template.[/red]")
        return

    # Read rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        def get(col):
            i = idx.get(col)
            return str(row[i]).strip() if i is not None and row[i] is not None else ""
        if not get("summary") or get("summary").startswith("📌"):
            continue
        rows.append({
            "summary":        get("summary"),
            "parent_issue":   get("parent_issue"),
            "description":    get("description"),
            "priority":       get("priority") or "Medium",
            "story_points":   get("story_points"),
            "assignee_email": get("assignee_email"),
        })

    if not rows:
        console.print("[yellow]⚠️ No rows found in the sheet.[/yellow]")
        return

    console.print(f"\n[cyan]Found {len(rows)} row(s) to create:[/cyan]")
    table = Table(show_header=True, header_style="bold cyan")
    for col in ["#", "Summary", "Parent", "Priority", "Story Points"]:
        table.add_column(col)
    for i, r in enumerate(rows, 1):
        table.add_row(str(i), r["summary"][:50], r["parent_issue"], r["priority"], r["story_points"])
    console.print(table)

    confirm = inquirer.prompt([
        inquirer.Confirm("go", message=f"Create {len(rows)} stories in Jira?", default=True)
    ])
    if not confirm or not confirm["go"]:
        console.print("[yellow]Cancelled.[/yellow]")
        return

    sp_field = get_story_points_field(email, token)
    created, failed = [], []

    for i, story in enumerate(rows, 1):
        console.print(f"[dim]Creating {i}/{len(rows)}: {story['summary'][:60]}...[/dim]")

        payload = {
            "fields": {
                "project":     {"key": JIRA_PROJECT},
                "summary":     story["summary"],
                "issuetype":   {"name": "Story"},
                "priority":    {"name": story["priority"]},
                "description": {
                    "type": "doc", "version": 1,
                    "content": [{"type": "paragraph", "content": [{"type": "text", "text": story["description"] or " "}]}],
                },
            }
        }

        sp_raw = story["story_points"].lower().strip()
        sp_val = STORY_POINTS_MAP.get(sp_raw)
        if sp_val and sp_field:
            payload["fields"][sp_field] = float(sp_val)

        if story["parent_issue"]:
            payload["fields"]["parent"] = {"key": story["parent_issue"].upper()}

        if story["assignee_email"]:
            r = requests.get(
                f"{JIRA_BASE_URL}/rest/api/3/user/search",
                params={"query": story["assignee_email"]},
                auth=auth(email, token),
                headers={"Accept": "application/json"},
                timeout=10,
            )
            if r.status_code == 200 and r.json():
                payload["fields"]["assignee"] = {"accountId": r.json()[0]["accountId"]}

        resp = requests.post(
            f"{JIRA_BASE_URL}/rest/api/3/issue",
            auth=auth(email, token),
            headers={"Accept": "application/json", "Content-Type": "application/json"},
            json=payload,
            timeout=15,
        )

        if resp.status_code == 201:
            data = resp.json()
            url  = f"{JIRA_BASE_URL}/browse/{data['key']}"
            created.append({"row": i, "key": data["key"], "summary": story["summary"], "url": url})
            console.print(f"  [green]✅ {data['key']}[/green] — {story['summary'][:50]}")
        else:
            failed.append({"row": i, "summary": story["summary"], "error": resp.text})
            console.print(f"  [red]❌ Row {i} failed:[/red] {resp.text}")

    # Write results back to Excel
    if created:
        console.print(f"\n[green]✅ {len(created)} ticket(s) created. Writing keys back to Excel...[/green]")
        # Add Jira Key + URL columns if not present
        key_col   = ws.max_column + 1
        url_col   = ws.max_column + 2
        ws.cell(row=1, column=key_col, value="Jira Key").font   = Font(bold=True)
        ws.cell(row=1, column=url_col, value="Jira URL").font   = Font(bold=True)

        for item in created:
            ws.cell(row=item["row"] + 1, column=key_col, value=item["key"])
            ws.cell(row=item["row"] + 1, column=url_col, value=item["url"])

        wb.save(filepath)
        console.print(f"[green]✅ Jira keys written back to:[/green] {filepath}")

    if failed:
        console.print(f"[red]❌ {len(failed)} row(s) failed.[/red]")


# ── Sync Jira Data to Excel ────────────────────────────────────────────────────

def sync_jira_to_excel(email, token):
    console.rule("[bold blue]📤 Sync Jira Data to Excel[/bold blue]")

    today = datetime.today().date()
    ans = inquirer.prompt([
        inquirer.List("board",    message="Select Board", choices=list(BOARDS.keys())),
        inquirer.Text("start",    message="From date (YYYY-MM-DD)", default=str(today - timedelta(days=30))),
        inquirer.Text("end",      message="To date (YYYY-MM-DD)",   default=str(today)),
        inquirer.Checkbox("types", message="Issue Types",
                          choices=["Story", "Bug", "Task", "Epic", "Sub-task"],
                          default=["Story", "Bug", "Task"]),
        inquirer.Checkbox("statuses", message="Statuses (leave empty for all)",
                          choices=["To Do", "In Progress", "In Review", "Done", "Blocked"],
                          default=[]),
        inquirer.Text("parent",   message="Filter by Parent Issue (optional)", default=""),
        inquirer.Text("max",      message="Max results", default="100"),
        inquirer.Confirm("comments", message="Also pull comments for each issue?", default=False),
        inquirer.Text("file",     message="Save to Excel file", default=f"jira_sync_{today}.xlsx"),
    ])

    type_clause   = " OR ".join([f'issuetype = "{t}"' for t in ans["types"]]) if ans["types"] else 'issuetype in standardIssueTypes()'
    status_clause = (" AND (" + " OR ".join([f'status = "{s}"' for s in ans["statuses"]]) + ")") if ans["statuses"] else ""
    parent_clause = f' AND parent = {ans["parent"].strip().upper()}' if ans["parent"].strip() else ""
    date_clause   = f' AND created >= "{ans["start"]}" AND created <= "{ans["end"]}"' if not ans["parent"].strip() else ""
    jql = f'project = {JIRA_PROJECT} AND ({type_clause}){date_clause}{status_clause}{parent_clause} ORDER BY created DESC'

    console.print(f"\n[dim]JQL: {jql}[/dim]")

    with console.status("Fetching data from Jira..."):
        resp = requests.post(
            f"{JIRA_BASE_URL}/rest/api/3/search/jql",
            auth=auth(email, token),
            headers={"Accept": "application/json", "Content-Type": "application/json"},
            json={
                "jql": jql,
                "maxResults": int(ans["max"]),
                "fields": ["summary", "status", "issuetype", "priority", "assignee",
                           "reporter", "created", "updated", "labels", "components", "parent"],
            },
            timeout=30,
        )

    if resp.status_code != 200:
        console.print(f"[red]❌ Failed: {resp.status_code}\n{resp.text}[/red]")
        return

    raw    = resp.json()
    issues = []
    for issue in raw.get("issues", []):
        f = issue.get("fields", {})
        issues.append({
            "Key":        issue["key"],
            "Summary":    f.get("summary", ""),
            "Type":       f.get("issuetype", {}).get("name", ""),
            "Status":     f.get("status", {}).get("name", ""),
            "Priority":   f.get("priority", {}).get("name", ""),
            "Parent":     f.get("parent", {}).get("key", "") if f.get("parent") else "",
            "Assignee":   f.get("assignee", {}).get("displayName", "") if f.get("assignee") else "",
            "Reporter":   f.get("reporter", {}).get("displayName", "") if f.get("reporter") else "",
            "Labels":     ", ".join(f.get("labels", [])),
            "Components": ", ".join(c["name"] for c in f.get("components", [])),
            "Created":    (f.get("created") or "")[:10],
            "Updated":    (f.get("updated") or "")[:10],
            "URL":        f"{JIRA_BASE_URL}/browse/{issue['key']}",
        })

    console.print(f"\n[green]✅ Fetched {len(issues)} of {raw.get('total', 0)} issues.[/green]")

    if not issues:
        console.print("[yellow]No issues to write.[/yellow]")
        return

    comments = []
    if ans["comments"]:
        with console.status("Fetching comments..."):
            for issue in issues:
                comments.extend(fetch_issue_comments(email, token, issue["Key"]))
        console.print(f"[green]✅ Fetched {len(comments)} comment(s) across {len(issues)} issue(s).[/green]")

    # Write to Excel
    filename = ans["file"]
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    # Load existing or create new workbook
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    sheet_name = f"Jira Sync {ans['start']}"[:31]
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    headers = list(issues[0].keys())
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill  = header_fill
        cell.font  = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    row_fill_even = PatternFill("solid", fgColor="EBF3FB")
    for r_idx, issue in enumerate(issues, 2):
        for c_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=issue[key])
            if r_idx % 2 == 0:
                cell.fill = row_fill_even

    if ans["comments"]:
        comments_sheet_name = f"Comments {ans['start']}"[:31]
        if comments_sheet_name in wb.sheetnames:
            del wb[comments_sheet_name]
        cws = wb.create_sheet(title=comments_sheet_name)

        c_headers = ["Key", "Author", "Created", "Text"]
        for col, h in enumerate(c_headers, 1):
            cell = cws.cell(row=1, column=col, value=h)
            cell.fill  = header_fill
            cell.font  = header_font
            cell.alignment = Alignment(horizontal="center")
            cws.column_dimensions[cell.column_letter].width = 20 if h != "Text" else 60

        for r_idx, comment in enumerate(comments, 2):
            for c_idx, key in enumerate(c_headers, 1):
                cell = cws.cell(row=r_idx, column=c_idx, value=comment[key])
                if r_idx % 2 == 0:
                    cell.fill = row_fill_even

    wb.save(filename)
    console.print(f"[green]✅ Jira data saved to:[/green] {filename} (sheet: '{sheet_name}')")
    if ans["comments"]:
        console.print(f"[green]✅ Comments saved to sheet:[/green] '{comments_sheet_name}'")


# ── Main Menu ──────────────────────────────────────────────────────────────────

def main():
    console.print("\n[bold cyan]📊 Excel ↔ Jira CLI[/bold cyan]")
    console.print(f"[dim]Project: {JIRA_PROJECT} | {JIRA_BASE_URL}[/dim]\n")

    email, token = get_credentials()
    if not validate_credentials(email, token):
        sys.exit(1)

    while True:
        ans = inquirer.prompt([
            inquirer.List("action", message="What would you like to do?", choices=[
                "📄 Generate Excel Template",
                "📥 Bulk Create Stories from Excel",
                "📤 Sync Jira Data to Excel",
                "🚪 Exit",
            ])
        ])
        choice = ans["action"]
        if "Template"   in choice:
            generate_template()
        elif "Create"   in choice:
            bulk_create_from_excel(email, token)
        elif "Sync"     in choice:
            sync_jira_to_excel(email, token)
        elif "Exit"     in choice:
            console.print("[cyan]Bye! 👋[/cyan]")
            break


if __name__ == "__main__":
    main()
